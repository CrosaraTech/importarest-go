"""Migrate RELACAO_EMPRESAS XLSX into the Supabase companies table.

Usage:
    python scripts/migrate_companies.py --xlsx path/to/RELACAO_EMPRESAS.xlsx

Upserts all rows (no municipality filter) into the Supabase companies table.
Safe to run multiple times — uses on_conflict="cod" so existing rows are updated
rather than duplicated.

IMPORTANT: Never imports from config.py (which has G: drive paths).
All Supabase credentials come from environment variables (loaded via python-dotenv).
"""
import argparse
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client

load_dotenv()

# ---------------------------------------------------------------------------
# Column index constants (confirmed from config.py source of truth)
# These columns are at fixed positions in every version of RELACAO_EMPRESAS.xlsx.
# ---------------------------------------------------------------------------
_COL_COD = 0       # A — company code (upsert key)
_COL_RAZAO = 1     # B — Razao Social
_COL_ANALISTA = 3  # D — analyst name (must match profiles.analyst_name exactly)
_COL_IM = 7        # H — Inscricao Municipal

# CNPJ, MUNICIPIO, and NOME EMPRESA positions vary across XLSX versions,
# so we discover them dynamically by header name (safer per RESEARCH Open Question 2).
_DYNAMIC_COLS = {"CNPJ", "MUNICIPIO", "NOME EMPRESA"}


def migrate(xlsx_path: Path) -> int:
    """Read all rows from xlsx_path and upsert into Supabase companies table.

    Returns the number of records upserted.
    """
    supabase = create_client(
        os.environ["SUPABASE_URL"],
        os.environ["SUPABASE_SECRET_KEY"],
    )

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # Build header map: {HEADER_NAME_UPPER: column_index}
    header_row = next(rows_iter)
    header_map: dict[str, int] = {
        str(c).strip().upper(): i
        for i, c in enumerate(header_row)
        if c is not None and str(c).strip()
    }

    def _get(row: tuple, index: int | None) -> str:
        """Safely extract a string value from a row tuple by index."""
        if index is None:
            return ""
        if index >= len(row):
            return ""
        val = row[index]
        if val is None:
            return ""
        return str(val).strip()

    records = []
    for row in rows_iter:
        cod = _get(row, _COL_COD)
        if not cod:
            # Skip rows with no COD — these are blank spacer rows in the XLSX
            continue

        records.append({
            "cod": cod,
            "razao": _get(row, _COL_RAZAO),
            "analista": _get(row, _COL_ANALISTA),
            "im": _get(row, _COL_IM),
            "municipio": _get(row, header_map.get("MUNICIPIO")),
            "cnpj": _get(row, header_map.get("CNPJ")),
            "nome_empresa": _get(row, header_map.get("NOME EMPRESA")),
        })

    wb.close()

    if not records:
        print("No records found to upsert.")
        return 0

    # Single-batch upsert (368 rows — well within Supabase insert limit of 500MB/request)
    result = supabase.table("companies").upsert(records, on_conflict="cod").execute()
    count = len(result.data) if result.data else len(records)
    print(f"Upserted {count} rows into companies table.")
    return count


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Migrate RELACAO_EMPRESAS XLSX into the Supabase companies table."
    )
    parser.add_argument(
        "--xlsx",
        required=True,
        type=Path,
        help="Path to the RELACAO_EMPRESAS.xlsx file",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"ERROR: File not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    migrate(args.xlsx)


if __name__ == "__main__":
    main()
