"""
services/spreadsheet.py — XLSX reader para planilha de empresas.

Funções públicas:
    load_analysts()                  -> list[str]
    get_companies_for_analyst(str)   -> list[dict]

Exceções:
    SpreadsheetError         (base)
    SpreadsheetAccessError   (arquivo não encontrado, bloqueado ou inacessível)
    SpreadsheetFormatError   (cabeçalho ausente, arquivo corrompido)
"""
from __future__ import annotations

import zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from config import (
    PLANILHA_EMPRESAS, PLANILHA_COL_COD, PLANILHA_COL_ANALISTA,
    PLANILHA_COL_IM, PLANILHA_COL_RAZAO, MUNICIPIOS_ACEITOS,
)


# ---------------------------------------------------------------------------
# Exceções públicas
# ---------------------------------------------------------------------------

class SpreadsheetError(Exception):
    """Exceção base para todas as falhas relacionadas à planilha."""


class SpreadsheetAccessError(SpreadsheetError):
    """Arquivo não encontrado, bloqueado ou inacessível."""


class SpreadsheetFormatError(SpreadsheetError):
    """Arquivo encontrado mas com cabeçalhos ausentes ou corrompido."""


# ---------------------------------------------------------------------------
# Implementação interna
# ---------------------------------------------------------------------------

def _load_goiania_rows() -> list[dict]:
    """Abre a planilha, valida cabeçalhos e retorna linhas filtradas para municípios aceitos.

    Returns:
        list[dict]: Cada item contém {"cod": str, "analista": str}.

    Raises:
        SpreadsheetAccessError: Se o arquivo não existe ou está bloqueado.
        SpreadsheetFormatError: Se o cabeçalho está ausente/incorreto ou arquivo corrompido.
    """
    path = PLANILHA_EMPRESAS
    wb = None
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        # Lê a primeira linha como cabeçalho
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise SpreadsheetFormatError(
                f"Planilha vazia — nenhuma linha de cabeçalho encontrada em: {path}"
            )

        # Monta mapa de nome_coluna -> índice
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is not None:
                header_map[str(cell).strip().upper()] = idx

        # Valida coluna COD (deve estar na posição PLANILHA_COL_COD = 0, coluna A)
        if header_map.get("COD") != PLANILHA_COL_COD:
            raise SpreadsheetFormatError(
                f"Cabeçalho inválido: coluna COD não encontrada na coluna A "
                f"(índice {PLANILHA_COL_COD}). Cabeçalhos lidos: {list(header_map.keys())}"
            )

        # Valida coluna ANALISTA (deve estar na posição PLANILHA_COL_ANALISTA = 3, coluna D)
        if header_map.get("ANALISTA") != PLANILHA_COL_ANALISTA:
            raise SpreadsheetFormatError(
                f"Cabeçalho inválido: coluna ANALISTA não encontrada na coluna D "
                f"(índice {PLANILHA_COL_ANALISTA}). Cabeçalhos lidos: {list(header_map.keys())}"
            )

        # Localiza coluna MUNICIPIO dinamicamente (aceita acento ou sem acento)
        col_municipio = header_map.get("MUNICIPIO") or header_map.get("MUNICÍPIO")
        if col_municipio is None:
            raise SpreadsheetFormatError(
                f"Cabeçalho inválido: coluna MUNICIPIO não encontrada na planilha. "
                f"Cabeçalhos lidos: {list(header_map.keys())}"
            )

        _nomes_aceitos = {n.upper() for n in MUNICIPIOS_ACEITOS}

        filtered_rows: list[dict] = []
        for data_row in rows_iter:
            cod = data_row[PLANILHA_COL_COD] if len(data_row) > PLANILHA_COL_COD else None
            if not cod:
                continue
            municipio = data_row[col_municipio] if len(data_row) > col_municipio else None
            analista = data_row[PLANILHA_COL_ANALISTA] if len(data_row) > PLANILHA_COL_ANALISTA else None

            if str(municipio).strip().upper() in _nomes_aceitos:
                im = data_row[PLANILHA_COL_IM] if len(data_row) > PLANILHA_COL_IM else None
                razao = data_row[PLANILHA_COL_RAZAO] if len(data_row) > PLANILHA_COL_RAZAO else None
                filtered_rows.append({
                    "cod": str(cod).strip(),
                    "analista": str(analista).strip() if analista is not None else "",
                    "im": str(im).strip() if im is not None else "",
                    "razao": str(razao).strip() if razao is not None else "",
                })

        return filtered_rows

    except SpreadsheetError:
        raise

    except FileNotFoundError:
        raise SpreadsheetAccessError(
            f"Planilha não encontrada — verifique se o drive G: está conectado e o "
            f"arquivo existe em: {path}"
        )

    except PermissionError:
        raise SpreadsheetAccessError(
            f"A planilha está aberta ou bloqueada por outro processo. "
            f"Feche o arquivo e tente novamente: {path}"
        )

    except (zipfile.BadZipFile, InvalidFileException):
        raise SpreadsheetFormatError(
            f"O arquivo está corrompido ou inválido — não é um XLSX válido: {path}"
        )

    finally:
        if wb is not None:
            wb.close()


# ---------------------------------------------------------------------------
# API pública
# ---------------------------------------------------------------------------

def load_analysts() -> list[str]:
    """Retorna lista ordenada de nomes únicos de analistas com pelo menos uma empresa em GOIÂNIA.

    Returns:
        list[str]: Nomes de analistas ordenados alfabeticamente.

    Raises:
        SpreadsheetAccessError: Se o arquivo não pode ser lido.
        SpreadsheetFormatError: Se o arquivo tem formato inválido.
    """
    rows = _load_goiania_rows()
    analysts = sorted({row["analista"] for row in rows if row["analista"]})
    return analysts


def get_company_info(cod: str) -> dict | None:
    """Retorna dados da empresa pelo código (im, razao). None se não encontrada."""
    rows = _load_goiania_rows()
    for row in rows:
        if row["cod"] == cod.strip():
            return row
    return None


def get_companies_for_analyst(analista: str) -> list[dict]:
    """Retorna lista de empresas em GOIÂNIA para o analista especificado.

    Args:
        analista: Nome do analista (correspondência exata).

    Returns:
        list[dict]: Cada item contém {"cod": str, "analista": str}.
                    Chamador pode usar len() para obter a contagem.

    Raises:
        SpreadsheetAccessError: Se o arquivo não pode ser lido.
        SpreadsheetFormatError: Se o arquivo tem formato inválido.
    """
    rows = _load_goiania_rows()
    return [row for row in rows if row["analista"] == analista]
